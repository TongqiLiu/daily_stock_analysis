#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Options review workbook sync tool.

Capabilities:
1) Append transaction entries to `交易明细` with dedupe.
2) Keep row style consistent with existing historical rows.
3) Refresh `策略复盘` summary blocks (top stats / spread net / weekly expiry).

Input JSON format: see ../templates/options_updates.example.json
"""

from __future__ import annotations

import argparse
import json
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import load_workbook


COLS = 16


@dataclass
class Entry:
    open_date: str
    ticker: str
    action: str
    strike: Any
    expiry: str
    dte: Any = "-"
    open_stock_price: Any = "-"
    delta: Any = "-"
    iv: Any = "-"
    premium: Any = "-"
    close_date: str = ""
    pnl: Any = "-"
    pnl_pct: Any = "-"
    rule: str = "△"
    note: str = ""
    settle_full_premium: bool = False

    @staticmethod
    def from_dict(d: Dict[str, Any]) -> "Entry":
        return Entry(
            open_date=str(d.get("open_date", "-") or "-"),
            ticker=str(d.get("ticker", "") or "").strip(),
            action=str(d.get("action", "") or "").strip(),
            strike=d.get("strike", "-"),
            expiry=str(d.get("expiry", "") or "").strip(),
            dte=d.get("dte", "-"),
            open_stock_price=d.get("open_stock_price", "-"),
            delta=d.get("delta", "-"),
            iv=d.get("iv", "-"),
            premium=d.get("premium", "-"),
            close_date=str(d.get("close_date", "") or "").strip(),
            pnl=d.get("pnl", "-"),
            pnl_pct=d.get("pnl_pct", "-"),
            rule=str(d.get("rule", "△") or "△"),
            note=str(d.get("note", "") or ""),
            settle_full_premium=bool(d.get("settle_full_premium", False)),
        )


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Sync options review workbook from parsed image records")
    parser.add_argument("--workbook", required=True, help="Path to 期权记录_2026.xlsx")
    parser.add_argument("--input", required=True, help="Path to updates JSON")
    parser.add_argument("--dry-run", action="store_true", help="Only print planned actions, do not save")
    return parser.parse_args()


def as_float(val: Any) -> Optional[float]:
    if val is None:
        return None
    if isinstance(val, (int, float)):
        return float(val)
    s = str(val).strip()
    if not s or s == "-":
        return None
    s = s.replace("$", "").replace(",", "")
    if s.startswith("+"):
        s = s[1:]
    try:
        return float(s)
    except ValueError:
        return None


def money(v: float) -> str:
    return f"$+{abs(v):.2f}" if v >= 0 else f"$-{abs(v):.2f}"


def parse_money_cell(v: Any) -> Optional[float]:
    if isinstance(v, (int, float)):
        return float(v)
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    sign = -1.0 if "-" in s else 1.0
    digits = s.replace("$", "").replace("+", "").replace("-", "").replace(",", "").strip()
    try:
        return sign * float(digits)
    except ValueError:
        return None


def normalize_key_parts(*parts: Any) -> Tuple[str, ...]:
    return tuple(str(p if p is not None else "").strip() for p in parts)


def read_existing_dedupe_keys(ws) -> set:
    keys = set()
    for r in range(1, ws.max_row + 1):
        if not isinstance(ws.cell(r, 1).value, int):
            continue
        key = normalize_key_parts(
            ws.cell(r, 3).value,
            ws.cell(r, 4).value,
            ws.cell(r, 5).value,
            ws.cell(r, 6).value,
            ws.cell(r, 12).value,
            ws.cell(r, 13).value,
        )
        keys.add(key)
    return keys


def max_no(ws) -> int:
    m = 0
    for r in range(1, ws.max_row + 1):
        v = ws.cell(r, 1).value
        if isinstance(v, int):
            m = max(m, v)
    return m


def apply_full_premium_settlement(entry: Entry) -> None:
    """Apply '到期 OTM = 100%拿到权利金' shortcut for sold options."""
    prem = as_float(entry.premium)
    if not entry.action.startswith("卖") or prem is None:
        return
    entry.pnl = round(prem, 2)
    entry.pnl_pct = "100%"
    if not entry.close_date:
        entry.close_date = entry.expiry


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    for c in range(1, COLS + 1):
        ws.cell(dst_row, c)._style = copy(ws.cell(src_row, c)._style)


def append_entries(ws, payload: Dict[str, Any]) -> Tuple[int, int]:
    section_title = str(payload.get("section_title", "") or "").strip()
    entries = [Entry.from_dict(x) for x in payload.get("entries", [])]
    if not entries:
        return 0, 0

    existing = read_existing_dedupe_keys(ws)
    cur_no = max_no(ws)

    prepared: List[Tuple[Entry, Tuple[str, ...]]] = []
    skipped = 0
    for e in entries:
        if e.settle_full_premium:
            apply_full_premium_settlement(e)

        dedupe_key = normalize_key_parts(
            e.ticker,
            e.action,
            e.strike,
            e.expiry,
            e.close_date,
            e.pnl,
        )
        if dedupe_key in existing:
            skipped += 1
            continue
        prepared.append((e, dedupe_key))

    if not prepared:
        return 0, skipped

    ws.append([None] * COLS)
    ws.append([section_title] + [None] * (COLS - 1))
    header = [ws.cell(2, c).value for c in range(1, COLS + 1)]
    ws.append(header)

    title_row = ws.max_row - 1
    header_row = ws.max_row

    # style: section title use nearest previous section-title row style if present, fallback default
    for r in range(1, ws.max_row):
        v = ws.cell(r, 1).value
        if isinstance(v, str) and v.startswith("📅"):
            for c in range(1, COLS + 1):
                ws.cell(title_row, c)._style = copy(ws.cell(r, c)._style)
    copy_row_style(ws, 2, header_row)

    added = 0
    for e, dedupe_key in prepared:

        cur_no += 1
        row_vals = [
            cur_no,
            e.open_date,
            e.ticker,
            e.action,
            e.strike,
            e.expiry,
            e.dte,
            e.open_stock_price,
            e.delta,
            e.iv,
            e.premium,
            e.close_date,
            e.pnl,
            e.pnl_pct,
            e.rule,
            e.note,
        ]
        ws.append(row_vals)
        copy_row_style(ws, 62, ws.max_row)  # historical canonical data-row style

        existing.add(dedupe_key)
        added += 1

    return added, skipped


def collect_settled_rows(ws) -> List[Dict[str, Any]]:
    rows = []
    for r in range(1, ws.max_row + 1):
        no = ws.cell(r, 1).value
        if not isinstance(no, int):
            continue
        pnl = ws.cell(r, 13).value
        pnl_f = as_float(pnl)
        if pnl_f is None:
            continue
        rows.append(
            {
                "no": no,
                "ticker": str(ws.cell(r, 3).value or ""),
                "action": str(ws.cell(r, 4).value or ""),
                "strike": ws.cell(r, 5).value,
                "expiry": str(ws.cell(r, 6).value or ""),
                "dte": ws.cell(r, 7).value,
                "open": str(ws.cell(r, 2).value or ""),
                "close": str(ws.cell(r, 12).value or ""),
                "pnl": pnl_f,
                "rule": str(ws.cell(r, 15).value or ""),
                "note": str(ws.cell(r, 16).value or ""),
            }
        )
    return rows


def stats(group: List[Dict[str, Any]]) -> Tuple[int, int, int, str, float, float]:
    n = len(group)
    wins = sum(1 for x in group if x["pnl"] >= 0)
    losses = sum(1 for x in group if x["pnl"] < 0)
    win_rate = f"{round((wins / n) * 100):d}%" if n else "0%"
    total = sum(x["pnl"] for x in group)
    avg = total / n if n else 0.0
    return n, wins, losses, win_rate, total, avg


def refresh_top_summary(ws, settled_rows: List[Dict[str, Any]]) -> None:
    all_g = settled_rows
    seller = [x for x in all_g if x["action"].startswith("卖") and x["rule"] != "Spread"]
    buyer = [x for x in all_g if x["action"].startswith("买") and x["rule"] != "Spread"]
    spread = [x for x in all_g if x["rule"] == "Spread"]

    short = []
    for x in seller:
        try:
            dte = float(x["dte"])
        except Exception:
            dte = None
        if dte is not None and dte <= 7 and x["open"] == x["close"]:
            short.append(x)

    groups = [all_g, seller, buyer, spread, short]
    for idx, grp in enumerate(groups, start=3):
        n, w, l, wr, total, avg = stats(grp)
        ws.cell(idx, 2).value = n
        ws.cell(idx, 3).value = w
        ws.cell(idx, 4).value = l
        ws.cell(idx, 5).value = wr
        ws.cell(idx, 6).value = money(total)
        ws.cell(idx, 7).value = money(avg)


def find_row(ws, value: str) -> Optional[int]:
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == value:
            return r
    return None


def refresh_spread_section(ws, payload: Dict[str, Any]) -> None:
    updates = payload.get("spread_review_rows", [])
    if not updates:
        return

    header_row = find_row(ws, "Spread")
    if header_row is None:
        return

    # locate total row after header
    total_row = None
    r = header_row + 1
    while r <= ws.max_row:
        if ws.cell(r, 1).value == "合计":
            total_row = r
            break
        r += 1
    if total_row is None:
        return

    # collect existing rows and dedupe
    existing = {}
    for rr in range(header_row + 1, total_row):
        spread_name = ws.cell(rr, 1).value
        close_date = ws.cell(rr, 3).value
        net_cell = ws.cell(rr, 4).value
        if not spread_name:
            continue
        existing[(str(spread_name), str(close_date))] = {
            "spread": str(spread_name),
            "structure": ws.cell(rr, 2).value,
            "close_date": str(close_date),
            "net_pnl": parse_money_cell(net_cell) or 0.0,
            "result": ws.cell(rr, 5).value,
        }

    for item in updates:
        key = (str(item.get("spread", "")), str(item.get("close_date", "")))
        if not key[0]:
            continue
        existing[key] = {
            "spread": key[0],
            "structure": item.get("structure", ""),
            "close_date": key[1],
            "net_pnl": float(item.get("net_pnl", 0.0)),
            "result": item.get("result") or ("盈利" if float(item.get("net_pnl", 0.0)) >= 0 else "亏损"),
        }

    merged = sorted(existing.values(), key=lambda x: (x["close_date"], x["spread"]))

    # clear old data rows
    for rr in range(header_row + 1, total_row):
        for c in range(1, 6):
            ws.cell(rr, c).value = None

    # ensure enough rows by inserting before total row
    needed = len(merged)
    available = max(0, total_row - header_row - 1)
    if needed > available:
        insert_n = needed - available
        ws.insert_rows(total_row, insert_n)
        total_row += insert_n

    # write merged rows
    sample_style_row = header_row + 1
    for i, item in enumerate(merged):
        rr = header_row + 1 + i
        ws.cell(rr, 1).value = item["spread"]
        ws.cell(rr, 2).value = item["structure"]
        ws.cell(rr, 3).value = item["close_date"]
        ws.cell(rr, 4).value = money(item["net_pnl"])
        ws.cell(rr, 5).value = item["result"]
        copy_row_style(ws, sample_style_row, rr)

    # rewrite total
    total = sum(x["net_pnl"] for x in merged)
    ws.cell(total_row, 1).value = "合计"
    ws.cell(total_row, 2).value = f"{len(merged)} 个 Spread"
    ws.cell(total_row, 3).value = None
    ws.cell(total_row, 4).value = money(total)
    ws.cell(total_row, 5).value = "净盈利" if total >= 0 else "净亏损"

    # best-effort sync narrative row if it exists
    if ws.max_row >= 32 and ws.cell(32, 1).value == "🟢 优":
        wins = sum(1 for x in merged if x["net_pnl"] >= 0)
        losses = len(merged) - wins
        ws.cell(32, 3).value = f"{len(merged)} 个 Spread 净 {money(total)}，{wins} 盈 {losses} 亏"


def month_day_label(date_str: str) -> Optional[str]:
    try:
        dt = datetime.strptime(date_str, "%Y-%m-%d")
        return dt.strftime("%b %d")
    except Exception:
        return None


def refresh_weekly_section(ws, settled_rows: List[Dict[str, Any]], payload: Dict[str, Any]) -> None:
    header_row = None
    for r in range(1, ws.max_row + 1):
        if ws.cell(r, 1).value == "周" and ws.cell(r, 2).value == "笔数":
            header_row = r
            break
    if header_row is None:
        return

    total_row = None
    r = header_row + 1
    while r <= ws.max_row:
        if ws.cell(r, 1).value == "合计":
            total_row = r
            break
        r += 1
    if total_row is None:
        return

    # existing rows by month-day for overwrite
    existing_rows = {}
    for rr in range(header_row + 1, total_row):
        label = ws.cell(rr, 1).value
        if not isinstance(label, str):
            continue
        parts = label.split()
        if len(parts) >= 3:
            md = f"{parts[-2]} {parts[-1]}"
            existing_rows[md] = rr

    weekly_overrides = payload.get("weekly_note_overrides", {}) or {}
    weekly_exclude_ids = payload.get("weekly_exclude_ids", {}) or {}

    seller = [x for x in settled_rows if x["action"].startswith("卖") and x["rule"] != "Spread"]
    groups: Dict[str, List[Dict[str, Any]]] = {}
    for x in seller:
        exp = x["expiry"]
        try:
            datetime.strptime(exp, "%Y-%m-%d")
        except Exception:
            continue
        groups.setdefault(exp, []).append(x)

    exps = sorted(groups.keys())
    write_rows = []
    for exp in exps:
        grp = groups[exp]
        exclude_ids = {int(i) for i in weekly_exclude_ids.get(exp, [])}
        if exclude_ids:
            grp = [x for x in grp if int(x["no"]) not in exclude_ids]

        cnt = len(grp)
        itm = sum(1 for x in grp if ("被行权" in str(x["close"])) or ("被行权" in x["note"]) or abs(x["pnl"]) < 1e-9)
        otm = cnt - itm
        pnl = sum(x["pnl"] for x in grp)

        dt = datetime.strptime(exp, "%Y-%m-%d")
        label = f"W{dt.isocalendar().week} {dt.strftime('%b %d')}"
        default_note = "100% ✓" if cnt and itm == 0 else f"{round((otm / cnt) * 100) if cnt else 0:.0f}%"
        note = str(weekly_overrides.get(exp, default_note))
        write_rows.append((exp, label, cnt, otm, itm, pnl, note))

    # clear old rows
    for rr in range(header_row + 1, total_row):
        for c in range(1, 7):
            ws.cell(rr, c).value = None

    # ensure enough rows
    needed = len(write_rows)
    available = max(0, total_row - header_row - 1)
    if needed > available:
        ins = needed - available
        ws.insert_rows(total_row, ins)
        total_row += ins

    # write
    sample_style_row = header_row + 1
    sum_cnt = sum_otm = sum_itm = 0
    sum_pnl = 0.0
    for i, (_exp, label, cnt, otm, itm, pnl, note) in enumerate(write_rows):
        rr = header_row + 1 + i
        ws.cell(rr, 1).value = label
        ws.cell(rr, 2).value = cnt
        ws.cell(rr, 3).value = otm
        ws.cell(rr, 4).value = itm
        ws.cell(rr, 5).value = money(pnl)
        ws.cell(rr, 6).value = note
        copy_row_style(ws, sample_style_row, rr)

        sum_cnt += cnt
        sum_otm += otm
        sum_itm += itm
        sum_pnl += pnl

    ws.cell(total_row, 1).value = "合计"
    ws.cell(total_row, 2).value = sum_cnt
    ws.cell(total_row, 3).value = sum_otm
    ws.cell(total_row, 4).value = sum_itm
    ws.cell(total_row, 5).value = money(sum_pnl)
    ws.cell(total_row, 6).value = "—"

    # narrative sync row (best effort): row with "卖方到期策略稳定"
    for rr in range(1, ws.max_row + 1):
        if ws.cell(rr, 2).value == "卖方到期策略稳定":
            win_rate = round((sum_otm / sum_cnt) * 100) if sum_cnt else 0
            week_labels = [w[1].split()[0] for w in write_rows]
            ws.cell(rr, 3).value = f"{'/'.join(week_labels)} 共 {sum_cnt} 笔到期无价值，胜率 {win_rate}%（{sum_otm}/{sum_cnt}）"
            break


def refresh_strategy_sheet(ws_review, ws_trades, payload: Dict[str, Any]) -> None:
    settled_rows = collect_settled_rows(ws_trades)

    title_suffix = payload.get("summary_title_suffix")
    if isinstance(title_suffix, str) and title_suffix.strip():
        ws_review.cell(1, 1).value = f"📊 已结算腿数 P/L 总览（{title_suffix.strip()}）"

    refresh_top_summary(ws_review, settled_rows)
    refresh_spread_section(ws_review, payload)
    refresh_weekly_section(ws_review, settled_rows, payload)


def main() -> None:
    args = parse_args()
    workbook_path = Path(args.workbook).resolve()
    input_path = Path(args.input).resolve()

    payload = json.loads(input_path.read_text(encoding="utf-8"))

    wb = load_workbook(workbook_path)
    ws_trades = wb["交易明细"]
    ws_review = wb["策略复盘"]

    added, skipped = append_entries(ws_trades, payload)
    refresh_strategy_sheet(ws_review, ws_trades, payload)

    if args.dry_run:
        print(json.dumps({"added": added, "skipped": skipped, "saved": False}, ensure_ascii=False, indent=2))
        return

    wb.save(workbook_path)
    print(json.dumps({"added": added, "skipped": skipped, "saved": True, "workbook": str(workbook_path)}, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
